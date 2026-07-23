from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "import-templates"


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow


@dataclass(frozen=True)
class Col:
    key: str
    note: str = ""
    example: str = ""


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_table(ws, cols: list[Col], *, start_row: int = 1) -> None:
    header_row = start_row
    note_row = start_row + 1
    example_row = start_row + 2

    for i, c in enumerate(cols, start=1):
        h = ws.cell(row=header_row, column=i, value=c.key)
        h.fill = HEADER_FILL
        h.font = HEADER_FONT
        h.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

        n = ws.cell(row=note_row, column=i, value=c.note)
        n.fill = NOTE_FILL
        n.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

        ws.cell(row=example_row, column=i, value=c.example).alignment = Alignment(
            vertical="center", horizontal="left", wrap_text=True
        )

    ws.freeze_panes = ws["A4"]
    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[note_row].height = 48
    ws.row_dimensions[example_row].height = 20


def _autosize(ws, max_width: int = 42) -> None:
    widths: dict[int, float] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True):
        for idx, v in enumerate(row, start=1):
            s = "" if v is None else str(v)
            widths[idx] = max(widths.get(idx, 10.0), min(max_width, max(10.0, len(s) + 2)))
    _set_col_widths(ws, widths)


def _add_readme_sheet(wb: Workbook, title: str, lines: Iterable[str]) -> None:
    ws = wb.active
    ws.title = "README"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")
    r = 3
    for line in lines:
        ws.cell(row=r, column=1, value=line)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 120


def build_students_template(path: Path) -> None:
    wb = Workbook()
    _add_readme_sheet(
        wb,
        "Students Import Template",
        [
            "Fill the Students sheet first. Columns match the server Student schema.",
            "Dates must be in YYYY-MM-DD format (e.g., 2026-05-09).",
            "You can leave studentId empty; the server will auto-generate it if sessionId is provided.",
            "For lookup fields, prefer using the *Id columns* (sessionId, darjahId, subjectId, bookId, teacherId, gradeId, currentGradeId, previousGradeId).",
            "If you don't know an ID, leave it blank for now (import feature can be extended later to resolve by title/name).",
            "Optional multi-row data is in separate sheets (Guardians / PreviousSchools / LessonTrack) linked by studentId.",
        ],
    )

    ws = wb.create_sheet("Students")
    student_cols = [
        Col("sessionId", "MongoDB ObjectId for Session", ""),
        Col("studentId", "Leave blank to auto-generate (needs sessionId)", ""),
        Col("name.ur", "Required (Urdu name)", "طالب علم 1"),
        Col("name.en", "Optional (English name)", "Student 1"),
        Col("fatherName.ur", "Optional", "والد 1"),
        Col("fatherName.en", "Optional", "Father 1"),
        Col("gender", "male | female | blank", "male"),
        Col("idCard", "CNIC/B-Form (optional)", "35202-1234567-1"),
        Col("dateOfBirth", "YYYY-MM-DD (optional)", "2015-01-01"),
        Col("phone", "Optional", "+923001234567"),
        Col("country.ur", "Optional", "پاکستان"),
        Col("country.en", "Optional", "Pakistan"),
        Col("state.ur", "Optional", "پنجاب"),
        Col("state.en", "Optional", "Punjab"),
        Col("cityLoc.ur", "Optional", "لاہور"),
        Col("cityLoc.en", "Optional", "Lahore"),
        Col("districtCurrent.ur", "Optional", "لاہور"),
        Col("districtCurrent.en", "Optional", "Lahore"),
        Col("addressCurrent.ur", "Optional", "گلی نمبر 1"),
        Col("addressCurrent.en", "Optional", "Street 1"),
        Col("districtPermanent.ur", "Optional", "لاہور"),
        Col("districtPermanent.en", "Optional", "Lahore"),
        Col("addressPermanent.ur", "Optional", "گلی نمبر 1"),
        Col("addressPermanent.en", "Optional", "Street 1"),
        Col("classTypeLabel", "Optional (free text)", ""),
        Col("enrollmentDate", "YYYY-MM-DD (optional)", "2026-04-01"),
        Col("exitDate", "YYYY-MM-DD (optional)", ""),
        Col("exitReason.ur", "Optional", ""),
        Col("exitReason.en", "Optional", ""),
        Col("gradeId", "Legacy Grade ObjectId (optional)", ""),
        Col("currentGradeId", "Current Grade ObjectId (optional)", ""),
        Col("previousGradeId", "Previous Grade ObjectId (optional)", ""),
        Col("darjahId", "Darjah ObjectId (optional)", ""),
        Col("subjectId", "Subject ObjectId (optional)", ""),
        Col("bookId", "SubjectBook ObjectId (optional)", ""),
        Col("teacherId", "Teacher ObjectId (optional)", ""),
        Col("photoUrl", "Optional URL/path if you already have it", ""),
    ]
    _write_table(ws, student_cols)
    _autosize(ws)

    ws = wb.create_sheet("Guardians")
    guardian_cols = [
        Col("studentId", "Must match Students.studentId after import", ""),
        Col("name.ur", "Required for guardian row", "سرپرست 1"),
        Col("name.en", "Optional", "Guardian 1"),
        Col("relation.ur", "Optional", "والد"),
        Col("relation.en", "Optional", "Father"),
        Col("profession", "Optional", ""),
        Col("phone", "Optional", "+923001234567"),
        Col("idCard", "Optional", ""),
        Col("address.ur", "Optional", ""),
        Col("address.en", "Optional", ""),
    ]
    _write_table(ws, guardian_cols)
    _autosize(ws)

    ws = wb.create_sheet("PreviousSchools")
    prev_cols = [
        Col("studentId", "Must match Students.studentId after import", ""),
        Col("year", "e.g., 2024", "2024"),
        Col("grade", "free text", "5"),
        Col("institute", "free text", "Govt School"),
        Col("marks", "free text", "450"),
        Col("result", "free text", "Pass"),
    ]
    _write_table(ws, prev_cols)
    _autosize(ws)

    ws = wb.create_sheet("LessonTrack")
    lesson_cols = [
        Col("studentId", "Must match Students.studentId after import", ""),
        Col("para", "1..30 (required)", "1"),
        Col("startDate", "YYYY-MM-DD (optional)", "2026-04-10"),
        Col("endDate", "YYYY-MM-DD (optional)", ""),
    ]
    _write_table(ws, lesson_cols)
    _autosize(ws)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_students_sample(path: Path) -> None:
    wb = Workbook()
    _add_readme_sheet(
        wb,
        "Students Import Sample (Filled)",
        [
            "This file is a filled SAMPLE to show the exact format.",
            "Replace placeholder values like PUT_SESSION_ID_HERE with real MongoDB ObjectIds from your system.",
            "Keep dates as YYYY-MM-DD.",
            "studentId is left blank here to demonstrate auto-generation; it requires sessionId.",
            "IMPORTANT: Guardians/PreviousSchools/LessonTrack require studentId. If you want to import those too, first import Students, then copy the generated studentId values back into those sheets and re-import (or import those separately).",
        ],
    )

    ws = wb.create_sheet("Students")
    cols = [
        Col("sessionId", "MongoDB ObjectId for Session", "PUT_SESSION_ID_HERE"),
        Col("studentId", "Leave blank to auto-generate (needs sessionId)", ""),
        Col("name.ur", "Required (Urdu name)", "طالب علم 1"),
        Col("name.en", "Optional (English name)", "Student 1"),
        Col("fatherName.ur", "Optional", "والد 1"),
        Col("fatherName.en", "Optional", "Father 1"),
        Col("gender", "male | female | blank", "male"),
        Col("idCard", "CNIC/B-Form (optional)", "35202-1234567-1"),
        Col("dateOfBirth", "YYYY-MM-DD (optional)", "2015-01-01"),
        Col("phone", "Optional", "+923001234567"),
        Col("country.ur", "Optional", "پاکستان"),
        Col("country.en", "Optional", "Pakistan"),
        Col("state.ur", "Optional", "پنجاب"),
        Col("state.en", "Optional", "Punjab"),
        Col("cityLoc.ur", "Optional", "لاہور"),
        Col("cityLoc.en", "Optional", "Lahore"),
        Col("districtCurrent.ur", "Optional", "لاہور"),
        Col("districtCurrent.en", "Optional", "Lahore"),
        Col("addressCurrent.ur", "Optional", "گلی نمبر 1"),
        Col("addressCurrent.en", "Optional", "Street 1"),
        Col("districtPermanent.ur", "Optional", "لاہور"),
        Col("districtPermanent.en", "Optional", "Lahore"),
        Col("addressPermanent.ur", "Optional", "گلی نمبر 1"),
        Col("addressPermanent.en", "Optional", "Street 1"),
        Col("classTypeLabel", "Optional (free text)", "داخلہ"),
        Col("enrollmentDate", "YYYY-MM-DD (optional)", "2026-04-01"),
        Col("exitDate", "YYYY-MM-DD (optional)", ""),
        Col("exitReason.ur", "Optional", ""),
        Col("exitReason.en", "Optional", ""),
        Col("gradeId", "Legacy Grade ObjectId (optional)", ""),
        Col("currentGradeId", "Current Grade ObjectId (optional)", ""),
        Col("previousGradeId", "Previous Grade ObjectId (optional)", ""),
        Col("darjahId", "Darjah ObjectId (optional)", "PUT_DARJAH_ID_HERE"),
        Col("subjectId", "Subject ObjectId (optional)", "PUT_SUBJECT_ID_HERE"),
        Col("bookId", "SubjectBook ObjectId (optional)", "PUT_BOOK_ID_HERE"),
        Col("teacherId", "Teacher ObjectId (optional)", "PUT_TEACHER_ID_HERE"),
        Col("photoUrl", "Optional URL/path if you already have it", ""),
    ]
    _write_table(ws, cols)

    # Add 2 more sample rows (row 5 and 6) keeping same columns
    # Row 4 is the "example" row created by _write_table.
    rows = [
        {
            "sessionId": "PUT_SESSION_ID_HERE",
            "studentId": "",
            "name.ur": "طالب علم 2",
            "name.en": "Student 2",
            "fatherName.ur": "والد 2",
            "fatherName.en": "Father 2",
            "gender": "female",
            "idCard": "",
            "dateOfBirth": "2014-02-01",
            "phone": "+923111111111",
            "country.ur": "پاکستان",
            "country.en": "Pakistan",
            "state.ur": "سندھ",
            "state.en": "Sindh",
            "cityLoc.ur": "کراچی",
            "cityLoc.en": "Karachi",
            "districtCurrent.ur": "کراچی",
            "districtCurrent.en": "Karachi",
            "addressCurrent.ur": "گلی نمبر 2",
            "addressCurrent.en": "Street 2",
            "districtPermanent.ur": "کراچی",
            "districtPermanent.en": "Karachi",
            "addressPermanent.ur": "گلی نمبر 2",
            "addressPermanent.en": "Street 2",
            "classTypeLabel": "",
            "enrollmentDate": "2026-04-02",
            "darjahId": "PUT_DARJAH_ID_HERE",
            "subjectId": "PUT_SUBJECT_ID_HERE",
            "bookId": "",
            "teacherId": "",
        },
        {
            "sessionId": "PUT_SESSION_ID_HERE",
            "studentId": "",
            "name.ur": "طالب علم 3",
            "name.en": "Student 3",
            "fatherName.ur": "",
            "fatherName.en": "",
            "gender": "male",
            "idCard": "",
            "dateOfBirth": "",
            "phone": "",
            "country.ur": "پاکستان",
            "country.en": "Pakistan",
            "state.ur": "",
            "state.en": "",
            "cityLoc.ur": "",
            "cityLoc.en": "",
            "districtCurrent.ur": "",
            "districtCurrent.en": "",
            "addressCurrent.ur": "",
            "addressCurrent.en": "",
            "districtPermanent.ur": "",
            "districtPermanent.en": "",
            "addressPermanent.ur": "",
            "addressPermanent.en": "",
            "classTypeLabel": "",
            "enrollmentDate": "2026-04-03",
            "darjahId": "",
            "subjectId": "",
            "bookId": "",
            "teacherId": "",
        },
    ]
    key_to_col = {c.key: idx for idx, c in enumerate(cols, start=1)}
    for ridx, row in enumerate(rows, start=5):
        for k, v in row.items():
            ci = key_to_col.get(k)
            if ci is None:
                continue
            ws.cell(row=ridx, column=ci, value=v)

    _autosize(ws)

    ws = wb.create_sheet("Guardians")
    gcols = [
        Col("studentId", "Must match Students.studentId after import", "SAMPLE_STUDENT_ID_1"),
        Col("name.ur", "Required for guardian row", "سرپرست 1"),
        Col("name.en", "Optional", "Guardian 1"),
        Col("relation.ur", "Optional", "والد"),
        Col("relation.en", "Optional", "Father"),
        Col("profession", "Optional", "کاروبار"),
        Col("phone", "Optional", "+923001234567"),
        Col("idCard", "Optional", "35202-1234567-1"),
        Col("address.ur", "Optional", "پتہ"),
        Col("address.en", "Optional", "Address"),
    ]
    _write_table(ws, gcols)
    # extra guardian row
    ws.append(["SAMPLE_STUDENT_ID_2", "سرپرست 2", "Guardian 2", "والدہ", "Mother", "", "+923111111111", "", "", ""])
    _autosize(ws)

    ws = wb.create_sheet("PreviousSchools")
    pcols = [
        Col("studentId", "Must match Students.studentId after import", "SAMPLE_STUDENT_ID_1"),
        Col("year", "e.g., 2024", "2024"),
        Col("grade", "free text", "5"),
        Col("institute", "free text", "Govt School"),
        Col("marks", "free text", "450"),
        Col("result", "free text", "Pass"),
    ]
    _write_table(ws, pcols)
    ws.append(["SAMPLE_STUDENT_ID_2", "2023", "4", "Private School", "430", "Pass"])
    _autosize(ws)

    ws = wb.create_sheet("LessonTrack")
    lcols = [
        Col("studentId", "Must match Students.studentId after import", "SAMPLE_STUDENT_ID_1"),
        Col("para", "1..30 (required)", "1"),
        Col("startDate", "YYYY-MM-DD (optional)", "2026-04-10"),
        Col("endDate", "YYYY-MM-DD (optional)", ""),
    ]
    _write_table(ws, lcols)
    ws.append(["SAMPLE_STUDENT_ID_1", 2, "2026-04-15", ""])
    _autosize(ws)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_teachers_template(path: Path) -> None:
    wb = Workbook()
    _add_readme_sheet(
        wb,
        "Teachers Import Template",
        [
            "Fill the Teachers sheet first. Columns match the server Teacher schema.",
            "Dates must be in YYYY-MM-DD format (e.g., 2026-05-09).",
            "Teacher assignments (Session/Darjah/Subject/Book) are in the Assignments sheet linked by teacherKey.",
            "You can choose teacherKey as any unique value (e.g., T001) to link assignment rows.",
        ],
    )

    ws = wb.create_sheet("Teachers")
    teacher_cols = [
        Col("teacherKey", "Any unique key you choose to link assignments", "T001"),
        Col("name.ur", "Required", "استاد 1"),
        Col("name.en", "Optional", "Teacher 1"),
        Col("parentage.ur", "Optional", "والدیت"),
        Col("parentage.en", "Optional", "Parentage"),
        Col("idCard", "Optional", "35202-1234567-1"),
        Col("phone", "Optional", "+923001234567"),
        Col("maritalStatus", "single | married | widowed | divorced | blank", "married"),
        Col("dateOfBirth", "YYYY-MM-DD (optional)", "1990-01-01"),
        Col("country.ur", "Optional", "پاکستان"),
        Col("country.en", "Optional", "Pakistan"),
        Col("state.ur", "Optional", "پنجاب"),
        Col("state.en", "Optional", "Punjab"),
        Col("cityLoc.ur", "Optional", "لاہور"),
        Col("cityLoc.en", "Optional", "Lahore"),
        Col("districtCurrent.ur", "Optional", "لاہور"),
        Col("districtCurrent.en", "Optional", "Lahore"),
        Col("addressCurrent.ur", "Optional", ""),
        Col("addressCurrent.en", "Optional", ""),
        Col("districtPermanent.ur", "Optional", "لاہور"),
        Col("districtPermanent.en", "Optional", "Lahore"),
        Col("addressPermanent.ur", "Optional", ""),
        Col("addressPermanent.en", "Optional", ""),
        Col("deeniTaleem", "Optional", ""),
        Col("asriTaleem", "Optional", ""),
        Col("extraSkills", "Optional", ""),
        Col("jobStartDate", "YYYY-MM-DD (optional)", "2026-01-01"),
        Col("jobEndDate", "YYYY-MM-DD (optional)", ""),
        Col("status", "active | inactive | leave", "active"),
    ]
    _write_table(ws, teacher_cols)
    _autosize(ws)

    ws = wb.create_sheet("Assignments")
    ass_cols = [
        Col("teacherKey", "Must match Teachers.teacherKey", "T001"),
        Col("sessionId", "Required (MongoDB ObjectId for Session)", ""),
        Col("darjahId", "Optional (Darjah ObjectId)", ""),
        Col("subjectId", "Optional (Subject ObjectId)", ""),
        Col("bookId", "Optional (SubjectBook ObjectId)", ""),
    ]
    _write_table(ws, ass_cols)
    _autosize(ws)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_teachers_sample(path: Path) -> None:
    wb = Workbook()
    _add_readme_sheet(
        wb,
        "Teachers Import Sample (Filled)",
        [
            "This file is a filled SAMPLE to show the exact format.",
            "Replace placeholder values like PUT_SESSION_ID_HERE with real MongoDB ObjectIds from your system.",
            "Keep dates as YYYY-MM-DD.",
            "Assignments sheet rows are linked by teacherKey.",
        ],
    )

    ws = wb.create_sheet("Teachers")
    cols = [
        Col("teacherKey", "Any unique key you choose to link assignments", "T001"),
        Col("name.ur", "Required", "استاد 1"),
        Col("name.en", "Optional", "Teacher 1"),
        Col("parentage.ur", "Optional", "والدیت 1"),
        Col("parentage.en", "Optional", "Parentage 1"),
        Col("idCard", "Optional", "35202-1234567-1"),
        Col("phone", "Optional", "+923001234567"),
        Col("maritalStatus", "single | married | widowed | divorced | blank", "married"),
        Col("dateOfBirth", "YYYY-MM-DD (optional)", "1990-01-01"),
        Col("country.ur", "Optional", "پاکستان"),
        Col("country.en", "Optional", "Pakistan"),
        Col("state.ur", "Optional", "پنجاب"),
        Col("state.en", "Optional", "Punjab"),
        Col("cityLoc.ur", "Optional", "لاہور"),
        Col("cityLoc.en", "Optional", "Lahore"),
        Col("districtCurrent.ur", "Optional", "لاہور"),
        Col("districtCurrent.en", "Optional", "Lahore"),
        Col("addressCurrent.ur", "Optional", "گلی نمبر 1"),
        Col("addressCurrent.en", "Optional", "Street 1"),
        Col("districtPermanent.ur", "Optional", "لاہور"),
        Col("districtPermanent.en", "Optional", "Lahore"),
        Col("addressPermanent.ur", "Optional", "گلی نمبر 1"),
        Col("addressPermanent.en", "Optional", "Street 1"),
        Col("deeniTaleem", "Optional", "درس نظامی"),
        Col("asriTaleem", "Optional", "BA"),
        Col("extraSkills", "Optional", ""),
        Col("jobStartDate", "YYYY-MM-DD (optional)", "2026-01-01"),
        Col("jobEndDate", "YYYY-MM-DD (optional)", ""),
        Col("status", "active | inactive | leave", "active"),
    ]
    _write_table(ws, cols)
    # Add 2nd teacher row
    key_to_col = {c.key: idx for idx, c in enumerate(cols, start=1)}
    ws.insert_rows(5)
    row2 = {
        "teacherKey": "T002",
        "name.ur": "استاد 2",
        "name.en": "Teacher 2",
        "phone": "+923111111111",
        "maritalStatus": "single",
        "status": "active",
        "country.ur": "پاکستان",
        "country.en": "Pakistan",
        "state.ur": "سندھ",
        "state.en": "Sindh",
        "cityLoc.ur": "کراچی",
        "cityLoc.en": "Karachi",
    }
    for k, v in row2.items():
        ci = key_to_col.get(k)
        if ci:
            ws.cell(row=5, column=ci, value=v)
    _autosize(ws)

    ws = wb.create_sheet("Assignments")
    acols = [
        Col("teacherKey", "Must match Teachers.teacherKey", "T001"),
        Col("sessionId", "Required (MongoDB ObjectId for Session)", "PUT_SESSION_ID_HERE"),
        Col("darjahId", "Optional (Darjah ObjectId)", "PUT_DARJAH_ID_HERE"),
        Col("subjectId", "Optional (Subject ObjectId)", "PUT_SUBJECT_ID_HERE"),
        Col("bookId", "Optional (SubjectBook ObjectId)", "PUT_BOOK_ID_HERE"),
    ]
    _write_table(ws, acols)
    ws.append(["T001", "PUT_SESSION_ID_HERE", "PUT_DARJAH_ID_HERE", "PUT_SUBJECT_ID_HERE", ""])
    ws.append(["T002", "PUT_SESSION_ID_HERE", "PUT_DARJAH_ID_HERE", "PUT_SUBJECT_ID_HERE", "PUT_BOOK_ID_HERE"])
    _autosize(ws)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    build_students_template(OUT_DIR / "students-import-template.xlsx")
    build_teachers_template(OUT_DIR / "teachers-import-template.xlsx")
    build_students_sample(OUT_DIR / "students-import-sample.xlsx")
    build_teachers_sample(OUT_DIR / "teachers-import-sample.xlsx")
    print(f"Created templates in: {OUT_DIR}")


if __name__ == "__main__":
    main()

